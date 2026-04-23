"""Bulk import of the UK Road Safety dataset + proximity queries.

Previously lived in ``app/services/accident_import.py``; colocated here
because everything accident-related is part of the *risk* feature domain.
"""

from __future__ import annotations

import csv
import logging
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable, Iterator, Optional

from pymongo import InsertOne
from pymongo.errors import BulkWriteError

from app.core.config import get_settings
from app.core.types import GeoPoint
from app.db import ACCIDENTS_COLLECTION, get_collection
from app.risk.schemas import AccidentCreate, RoadSurface, Severity, WeatherCondition

logger = logging.getLogger(__name__)


_EXCEL_EPOCH = datetime(1899, 12, 30)


# ----------------------------- value parsers ----------------------------

def _parse_date(raw) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        return (_EXCEL_EPOCH + timedelta(days=float(raw))).date()
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return (_EXCEL_EPOCH + timedelta(days=float(s))).date()
    except ValueError:
        return None


def _parse_time(raw) -> Optional[time]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, time):
        return raw
    if isinstance(raw, datetime):
        return raw.time()
    if isinstance(raw, (int, float)):
        total = int(round(float(raw) * 24 * 60 * 60))
        total %= 24 * 60 * 60
        return time(total // 3600, (total % 3600) // 60, total % 60)
    s = str(raw).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    try:
        return _parse_time(float(s))
    except ValueError:
        return None


def _parse_float(raw) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _normalize_enum(raw, enum_cls, default):
    if raw is None or raw == "":
        return default
    s = str(raw).strip()
    for member in enum_cls:
        if member.value.lower() == s.lower():
            return member
    return default


def _row_to_accident(row: dict) -> Optional[AccidentCreate]:
    lat = _parse_float(row.get("Latitude"))
    lon = _parse_float(row.get("Longitude"))
    dt = _parse_date(row.get("Date"))
    if lat is None or lon is None or dt is None:
        return None
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return None

    return AccidentCreate(
        severity=_normalize_enum(row.get("Severity"), Severity, Severity.slight),
        date=dt,
        day_of_week=str(row.get("Day_of_Week") or "").strip() or "Unknown",
        latitude=lat,
        longitude=lon,
        road_surface=_normalize_enum(
            row.get("Road_Surface"), RoadSurface, RoadSurface.unknown
        ),
        time=_parse_time(row.get("Time")),
        weather_conditions=_normalize_enum(
            row.get("Weather_Conditions"), WeatherCondition, WeatherCondition.unknown
        ),
    )


# ------------------------------- sources --------------------------------

def _iter_csv_rows(path: Path) -> Iterator[dict]:
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def _iter_xlsx_rows(path: Path) -> Iterator[dict]:
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        header: Optional[list] = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            if all(v is None for v in row):
                continue
            yield dict(zip(header, row))
    finally:
        wb.close()


def _iter_dataset_rows(path: Path) -> Iterator[dict]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _iter_xlsx_rows(path)
    if suffix == ".csv":
        return _iter_csv_rows(path)
    return _iter_xlsx_rows(path)


def _chunks(it: Iterable, size: int) -> Iterator[list]:
    buf: list = []
    for item in it:
        buf.append(item)
        if len(buf) >= size:
            yield buf
            buf = []
    if buf:
        yield buf


# ------------------------------- bulk op --------------------------------

async def import_accidents_from_file(path: Path, batch_size: int = 5000) -> int:
    if not path.exists():
        logger.warning("Accident dataset not found at %s - skipping import.", path)
        return 0

    collection = get_collection(ACCIDENTS_COLLECTION)
    logger.info("Importing accidents from %s (batch size %d)...", path, batch_size)

    inserted_total = 0
    skipped_total = 0

    def build_ops() -> Iterator[InsertOne]:
        nonlocal skipped_total
        for raw in _iter_dataset_rows(path):
            accident = _row_to_accident(raw)
            if accident is None:
                skipped_total += 1
                continue
            yield InsertOne(accident.to_document())

    for batch in _chunks(build_ops(), batch_size):
        try:
            result = await collection.bulk_write(batch, ordered=False)
            inserted_total += result.inserted_count
        except BulkWriteError as bwe:
            inserted_total += bwe.details.get("nInserted", 0)
            logger.warning(
                "Partial bulk write: %d errors in this batch",
                len(bwe.details.get("writeErrors", [])),
            )
        if inserted_total and inserted_total % (batch_size * 10) == 0:
            logger.info("  ...%d accidents inserted so far", inserted_total)

    logger.info(
        "Accident import finished: inserted=%d skipped=%d",
        inserted_total,
        skipped_total,
    )
    return inserted_total


async def import_accidents_if_empty() -> int:
    """Seed accident_history from the dataset when empty. Safe on every boot."""
    settings = get_settings()
    if not settings.import_accidents_on_startup:
        logger.info("IMPORT_ACCIDENTS_ON_STARTUP=false - skipping dataset import.")
        return 0

    collection = get_collection(ACCIDENTS_COLLECTION)
    existing = await collection.estimated_document_count()
    if existing > 0:
        logger.info(
            "accident_history already has ~%d docs - skipping dataset import.",
            existing,
        )
        return 0

    return await import_accidents_from_file(
        settings.dataset_absolute_path,
        batch_size=settings.accidents_import_batch_size,
    )


# --------------------------- proximity queries --------------------------

async def find_accidents_within_radius(
    latitude: float,
    longitude: float,
    radius_meters: float = 500.0,
    limit: int = 100,
) -> list[dict]:
    collection = get_collection(ACCIDENTS_COLLECTION)
    cursor = collection.find(
        {
            "location": {
                "$nearSphere": {
                    "$geometry": GeoPoint.from_lat_lon(latitude, longitude).model_dump(),
                    "$maxDistance": radius_meters,
                }
            }
        }
    ).limit(limit)
    return [doc async for doc in cursor]


async def count_accidents_within_radius(
    latitude: float,
    longitude: float,
    radius_meters: float = 500.0,
) -> int:
    """Exact count of accidents inside the radius (used as the h_loc_count feature)."""
    collection = get_collection(ACCIDENTS_COLLECTION)
    return await collection.count_documents(
        {
            "location": {
                "$geoWithin": {
                    "$centerSphere": [
                        [longitude, latitude],
                        radius_meters / 6_378_100.0,  # meters -> radians
                    ]
                }
            }
        }
    )
